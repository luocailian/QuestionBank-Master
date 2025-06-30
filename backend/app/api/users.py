"""
用户管理API
"""
from flask import request
from flask_restx import Namespace, Resource, fields
from flask_jwt_extended import jwt_required, get_jwt_identity, get_jwt
from marshmallow import Schema, fields as ma_fields, validate, ValidationError
from datetime import datetime

from app import db
from app.models import User, UserProgress, UserPoints

# 创建命名空间
users_bp = Namespace('users', description='用户管理相关接口')

# 响应模型
user_model = users_bp.model('User', {
    'id': fields.Integer(description='用户ID'),
    'username': fields.String(description='用户名'),
    'email': fields.String(description='邮箱'),
    'role': fields.String(description='角色'),
    'avatar_url': fields.String(description='头像URL'),
    'created_at': fields.String(description='创建时间'),
    'is_active': fields.Boolean(description='是否激活')
})

profile_update_model = users_bp.model('ProfileUpdate', {
    'username': fields.String(description='用户名'),
    'email': fields.String(description='邮箱'),
    'avatar_url': fields.String(description='头像URL')
})

password_change_model = users_bp.model('PasswordChange', {
    'old_password': fields.String(required=True, description='旧密码'),
    'new_password': fields.String(required=True, description='新密码')
})

# Marshmallow验证模式
class ProfileUpdateSchema(Schema):
    username = ma_fields.Str(validate=validate.Length(min=3, max=50))
    email = ma_fields.Email()
    avatar_url = ma_fields.Url()

class PasswordChangeSchema(Schema):
    old_password = ma_fields.Str(required=True, validate=validate.Length(min=6, max=128))
    new_password = ma_fields.Str(required=True, validate=validate.Length(min=6, max=128))

def admin_required():
    """检查是否为管理员"""
    claims = get_jwt()
    if claims.get('role') != 'admin':
        return {'message': '需要管理员权限'}, 403
    return None

@users_bp.route('')
class UserList(Resource):
    @jwt_required()
    def get(self):
        """获取用户列表（管理员专用）"""
        # 检查管理员权限
        admin_check = admin_required()
        if admin_check:
            return admin_check

        # 获取查询参数
        page = int(request.args.get('page', 1))
        per_page = min(int(request.args.get('per_page', 20)), 100)
        search = request.args.get('search', '')
        role = request.args.get('role', '')

        # 构建查询
        query = User.query

        # 搜索过滤
        if search:
            query = query.filter(
                db.or_(
                    User.username.contains(search),
                    User.email.contains(search),
                    User.real_name.contains(search)
                )
            )

        # 角色过滤
        if role:
            query = query.filter(User.role == role)

        # 分页查询
        pagination = query.order_by(User.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )

        users = []
        for user in pagination.items:
            user_data = user.to_dict()
            # 添加统计信息
            stats = user.get_statistics()
            user_data['statistics'] = stats
            users.append(user_data)

        return {
            'users': users,
            'total': pagination.total,
            'page': page,
            'per_page': per_page,
            'pages': pagination.pages
        }

@users_bp.route('/<int:user_id>')
class UserDetail(Resource):
    @jwt_required()
    def get(self, user_id):
        """获取用户详情"""
        # 检查权限：管理员或用户本人
        current_user_id = get_jwt_identity()
        claims = get_jwt()

        if claims.get('role') != 'admin' and int(current_user_id) != user_id:
            return {'message': '权限不足'}, 403

        user = User.query.get_or_404(user_id)
        user_data = user.to_dict()

        # 添加统计信息
        stats = user.get_statistics()
        user_data['statistics'] = stats

        return user_data

    @jwt_required()
    def put(self, user_id):
        """更新用户信息"""
        # 检查管理员权限
        admin_check = admin_required()
        if admin_check:
            return admin_check

        user = User.query.get_or_404(user_id)
        data = request.get_json()

        # 更新允许的字段
        if 'username' in data:
            # 检查用户名是否已存在
            existing = User.query.filter(
                User.username == data['username'],
                User.id != user_id
            ).first()
            if existing:
                return {'message': '用户名已存在'}, 400
            user.username = data['username']

        if 'email' in data:
            # 检查邮箱是否已存在
            existing = User.query.filter(
                User.email == data['email'],
                User.id != user_id
            ).first()
            if existing:
                return {'message': '邮箱已存在'}, 400
            user.email = data['email']

        if 'role' in data:
            if data['role'] in ['admin', 'user']:
                user.role = data['role']

        if 'is_active' in data:
            user.is_active = bool(data['is_active'])

        try:
            db.session.commit()
            return user.to_dict()
        except Exception as e:
            db.session.rollback()
            return {'message': '更新失败'}, 500

    @jwt_required()
    def delete(self, user_id):
        """删除用户"""
        # 检查管理员权限
        admin_check = admin_required()
        if admin_check:
            return admin_check

        user = User.query.get_or_404(user_id)

        # 不能删除自己
        current_user_id = int(get_jwt_identity())
        if current_user_id == user_id:
            return {'message': '不能删除自己的账号'}, 400

        try:
            db.session.delete(user)
            db.session.commit()
            return {'message': '用户删除成功'}
        except Exception as e:
            db.session.rollback()
            return {'message': '删除失败'}, 500

@users_bp.route('/<int:user_id>/statistics')
class UserStatistics(Resource):
    @jwt_required()
    def get(self, user_id):
        """获取用户统计信息"""
        # 检查权限：管理员或用户本人
        current_user_id = get_jwt_identity()
        claims = get_jwt()

        if claims.get('role') != 'admin' and int(current_user_id) != user_id:
            return {'message': '权限不足'}, 403

        user = User.query.get_or_404(user_id)
        return user.get_statistics()

@users_bp.route('/admin/statistics')
class AdminStatistics(Resource):
    @jwt_required()
    def get(self):
        """获取管理后台统计数据"""
        # 检查管理员权限
        admin_check = admin_required()
        if admin_check:
            return admin_check

        try:
            # 获取用户统计
            total_users = User.query.count()
            active_users = User.query.filter_by(is_active=True).count()

            # 获取题库统计
            from app.models import QuestionBank, Question, UserAnswer
            total_banks = QuestionBank.query.count()

            # 获取题目统计
            total_questions = Question.query.count()

            # 获取答题统计
            total_answers = UserAnswer.query.count()
            correct_answers = UserAnswer.query.filter_by(is_correct=True).count()

            return {
                'total_users': total_users,
                'active_users': active_users,
                'total_banks': total_banks,
                'total_questions': total_questions,
                'total_answers': total_answers,
                'correct_answers': correct_answers,
                'accuracy_rate': round((correct_answers / total_answers * 100), 2) if total_answers > 0 else 0
            }
        except Exception as e:
            return {'message': f'获取统计数据失败: {str(e)}'}, 500

@users_bp.route('/export')
class UserExport(Resource):
    @jwt_required()
    def get(self):
        """导出用户数据"""
        # 检查管理员权限
        admin_check = admin_required()
        if admin_check:
            return admin_check

        # 获取查询参数
        search = request.args.get('search', '')
        role = request.args.get('role', '')
        status = request.args.get('status', '')

        # 构建查询
        query = User.query

        # 搜索过滤
        if search:
            query = query.filter(
                db.or_(
                    User.username.contains(search),
                    User.email.contains(search),
                    User.real_name.contains(search)
                )
            )

        # 角色过滤
        if role:
            query = query.filter(User.role == role)

        # 状态过滤
        if status:
            is_active = status == 'active'
            query = query.filter(User.is_active == is_active)

        users = query.order_by(User.created_at.desc()).all()

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            import io
            from flask import make_response

            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "用户列表"

            # 设置标题行
            headers = ['ID', '用户名', '邮箱', '角色', '状态', '注册时间', '最后登录', '创建题库数', '答题总数', '正确率', '总积分']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # 填充数据
            for row, user in enumerate(users, 2):
                stats = user.get_statistics()
                ws.cell(row=row, column=1, value=user.id)
                ws.cell(row=row, column=2, value=user.username)
                ws.cell(row=row, column=3, value=user.email)
                ws.cell(row=row, column=4, value='管理员' if user.role == 'admin' else '普通用户')
                ws.cell(row=row, column=5, value='正常' if user.is_active else '禁用')
                ws.cell(row=row, column=6, value=user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else '')
                ws.cell(row=row, column=7, value=user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else '从未登录')
                ws.cell(row=row, column=8, value=stats.get('total_banks', 0))
                ws.cell(row=row, column=9, value=stats.get('total_answers', 0))
                ws.cell(row=row, column=10, value=f"{stats.get('accuracy_rate', 0)}%")
                ws.cell(row=row, column=11, value=stats.get('total_points', 0))

            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # 创建响应
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename=users_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

            return response

        except ImportError:
            return {'message': 'openpyxl 库未安装，无法导出Excel格式'}, 500
        except Exception as e:
            return {'message': f'导出失败: {str(e)}'}, 500

@users_bp.route('/profile')
class UserProfile(Resource):
    @jwt_required()
    @users_bp.marshal_with(user_model)
    def get(self):
        """获取用户资料"""
        current_user_id = int(get_jwt_identity())  # 转换为整数
        user = User.query.get(current_user_id)
        
        if not user:
            return {'message': '用户不存在'}, 404
        
        return user.to_dict()
    
    @jwt_required()
    @users_bp.expect(profile_update_model)
    @users_bp.marshal_with(user_model)
    def put(self):
        """更新用户资料"""
        current_user_id = int(get_jwt_identity())  # 转换为整数
        user = User.query.get(current_user_id)
        
        if not user:
            return {'message': '用户不存在'}, 404
        
        try:
            # 验证请求数据
            schema = ProfileUpdateSchema()
            data = schema.load(request.json)
        except ValidationError as err:
            return {'message': '请求参数错误', 'errors': err.messages}, 400
        
        # 检查用户名是否已被其他用户使用
        if 'username' in data and data['username'] != user.username:
            existing_user = User.query.filter_by(username=data['username']).first()
            if existing_user:
                return {'message': '用户名已存在'}, 400
        
        # 检查邮箱是否已被其他用户使用
        if 'email' in data and data['email'] != user.email:
            existing_user = User.query.filter_by(email=data['email']).first()
            if existing_user:
                return {'message': '邮箱已存在'}, 400
        
        # 更新用户信息
        for key, value in data.items():
            setattr(user, key, value)
        
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return {'message': '更新失败，请稍后重试'}, 500
        
        return user.to_dict()

@users_bp.route('/change-password')
class ChangePassword(Resource):
    @jwt_required()
    @users_bp.expect(password_change_model)
    def post(self):
        """修改密码"""
        current_user_id = int(get_jwt_identity())  # 转换为整数
        user = User.query.get(current_user_id)
        
        if not user:
            return {'message': '用户不存在'}, 404
        
        try:
            # 验证请求数据
            schema = PasswordChangeSchema()
            data = schema.load(request.json)
        except ValidationError as err:
            return {'message': '请求参数错误', 'errors': err.messages}, 400
        
        # 验证旧密码
        if not user.check_password(data['old_password']):
            return {'message': '旧密码错误'}, 400
        
        # 设置新密码
        user.set_password(data['new_password'])
        
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return {'message': '密码修改失败，请稍后重试'}, 500
        
        return {'message': '密码修改成功'}

@users_bp.route('/statistics')
class UserStatistics(Resource):
    @jwt_required()
    def get(self):
        """获取用户统计信息"""
        current_user_id = int(get_jwt_identity())  # 转换为整数
        user = User.query.get(current_user_id)
        
        if not user:
            return {'message': '用户不存在'}, 404
        
        # 获取基础统计
        statistics = user.get_statistics()
        
        # 获取详细进度信息
        progress_list = UserProgress.query.filter_by(user_id=current_user_id).all()
        statistics['progress'] = [p.to_dict() for p in progress_list]
        
        # 获取积分信息
        user_points = UserPoints.query.filter_by(user_id=current_user_id).first()
        if user_points:
            statistics['points'] = user_points.to_dict()
        else:
            statistics['points'] = {
                'total_points': 0,
                'daily_points': 0,
                'weekly_points': 0,
                'monthly_points': 0
            }
        
        return statistics

@users_bp.route('/leaderboard')
class Leaderboard(Resource):
    def get(self):
        """获取排行榜"""
        # 获取查询参数
        period = request.args.get('period', 'total')  # total, daily, weekly, monthly
        limit = min(int(request.args.get('limit', 50)), 100)
        
        # 根据周期选择排序字段
        if period == 'daily':
            order_field = UserPoints.daily_points
        elif period == 'weekly':
            order_field = UserPoints.weekly_points
        elif period == 'monthly':
            order_field = UserPoints.monthly_points
        else:
            order_field = UserPoints.total_points
        
        # 查询排行榜
        leaderboard = db.session.query(User, UserPoints).join(
            UserPoints, User.id == UserPoints.user_id
        ).order_by(order_field.desc()).limit(limit).all()
        
        result = []
        for rank, (user, points) in enumerate(leaderboard, 1):
            result.append({
                'rank': rank,
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'avatar_url': user.avatar_url
                },
                'points': getattr(points, f'{period}_points' if period != 'total' else 'total_points')
            })
        
        return {
            'period': period,
            'leaderboard': result
        }
