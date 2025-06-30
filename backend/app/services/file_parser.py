"""
文件解析服务
支持PDF、DOCX、XLSX、JSON格式的题库文件解析
"""
import json
import re
from typing import List, Dict, Any

from app import db
from app.models import Question

class FileParserService:
    """文件解析服务类"""
    
    def parse_file(self, file_path: str, file_type: str) -> List[Dict[str, Any]]:
        """
        解析文件并返回题目数据
        
        Args:
            file_path: 文件路径
            file_type: 文件类型 (pdf, docx, xlsx, json)
            
        Returns:
            题目数据列表
        """
        if file_type == 'json':
            return self._parse_json(file_path)
        elif file_type == 'pdf':
            return self._parse_pdf(file_path)
        elif file_type == 'docx':
            return self._parse_docx(file_path)
        elif file_type == 'xlsx':
            return self._parse_xlsx(file_path)
        else:
            raise ValueError(f"不支持的文件类型: {file_type}")
    
    def _parse_json(self, file_path: str) -> List[Dict[str, Any]]:
        """解析JSON格式题库文件"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 验证JSON格式
            if isinstance(data, dict) and 'questions' in data:
                questions = data['questions']
            elif isinstance(data, list):
                questions = data
            else:
                raise ValueError("JSON格式不正确，应包含questions数组")
            
            # 验证每个题目的格式
            validated_questions = []
            for i, question in enumerate(questions):
                try:
                    validated_question = self._validate_question_data(question)
                    validated_question['order_index'] = i
                    validated_questions.append(validated_question)
                except Exception as e:
                    print(f"题目 {i+1} 格式错误: {e}")
                    continue
            
            return validated_questions
            
        except json.JSONDecodeError as e:
            raise ValueError(f"JSON文件格式错误: {e}")
        except Exception as e:
            raise ValueError(f"解析JSON文件失败: {e}")
    
    def _parse_pdf(self, file_path: str) -> List[Dict[str, Any]]:
        """解析PDF文件"""
        try:
            import fitz  # PyMuPDF
            
            doc = fitz.open(file_path)
            text = ""
            
            # 提取所有页面的文本
            for page in doc:
                text += page.get_text()
            
            doc.close()
            
            # 使用正则表达式解析题目
            return self._parse_text_content(text)
            
        except ImportError:
            raise ValueError("缺少PyMuPDF库，无法解析PDF文件")
        except Exception as e:
            raise ValueError(f"解析PDF文件失败: {e}")
    
    def _parse_docx(self, file_path: str) -> List[Dict[str, Any]]:
        """解析DOCX文件"""
        try:
            from docx import Document
            
            doc = Document(file_path)
            text = ""
            
            # 提取所有段落的文本
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            
            # 使用正则表达式解析题目
            return self._parse_text_content(text)
            
        except ImportError:
            raise ValueError("缺少python-docx库，无法解析DOCX文件")
        except Exception as e:
            raise ValueError(f"解析DOCX文件失败: {e}")
    
    def _parse_xlsx(self, file_path: str) -> List[Dict[str, Any]]:
        """解析XLSX文件"""
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            questions = []
            headers = []
            
            # 读取表头
            for cell in sheet[1]:
                headers.append(cell.value)
            
            # 读取数据行
            for row_num in range(2, sheet.max_row + 1):
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    cell_value = sheet.cell(row=row_num, column=col_num).value
                    if header:
                        row_data[header] = cell_value
                
                # 转换为标准格式
                if (row_data.get('题目') or row_data.get('题干') or
                    row_data.get('title') or row_data.get('question')):
                    question = self._convert_excel_row_to_question(row_data)
                    if question:
                        question['order_index'] = len(questions)
                        questions.append(question)
            
            workbook.close()
            return questions
            
        except ImportError:
            raise ValueError("缺少openpyxl库，无法解析XLSX文件")
        except Exception as e:
            raise ValueError(f"解析XLSX文件失败: {e}")
    
    def _parse_text_content(self, text: str) -> List[Dict[str, Any]]:
        """解析文本内容，提取题目"""
        questions = []
        
        # 简单的题目识别模式
        # 匹配形如 "1. 题目内容" 或 "题目1：题目内容" 的模式
        question_pattern = r'(?:^|\n)(?:题目)?(\d+)[.．：:]\s*(.+?)(?=(?:^|\n)(?:题目)?\d+[.．：:]|\Z)'
        
        matches = re.findall(question_pattern, text, re.MULTILINE | re.DOTALL)
        
        for i, (num, content) in enumerate(matches):
            content = content.strip()
            if not content:
                continue
            
            # 尝试识别题目类型和答案
            question = self._analyze_question_content(content)
            question['order_index'] = i
            questions.append(question)
        
        return questions
    
    def _analyze_question_content(self, content: str) -> Dict[str, Any]:
        """分析题目内容，识别题型和答案"""
        content = content.strip()
        
        # 检查是否为选择题
        if re.search(r'[ABCD][.．）)]\s*', content):
            return self._parse_choice_question(content)
        
        # 检查是否为判断题
        if re.search(r'[对错正误是否√×]', content) or '判断' in content:
            return self._parse_true_false_question(content)
        
        # 默认为问答题
        return self._parse_qa_question(content)
    
    def _parse_choice_question(self, content: str) -> Dict[str, Any]:
        """解析选择题"""
        lines = content.split('\n')
        title = lines[0].strip()
        
        options = []
        correct_option = None
        
        for line in lines[1:]:
            line = line.strip()
            if not line:
                continue
            
            # 匹配选项
            option_match = re.match(r'([ABCD])[.．）)]\s*(.+)', line)
            if option_match:
                option_key, option_text = option_match.groups()
                options.append({
                    'key': option_key,
                    'text': option_text.strip()
                })
            
            # 匹配答案
            answer_match = re.search(r'答案[：:]\s*([ABCD])', line)
            if answer_match:
                correct_option = answer_match.group(1)
        
        return {
            'type': 'choice',
            'title': title,
            'content': {
                'options': options
            },
            'answer': {
                'correct_option': correct_option or 'A'
            },
            'difficulty': 'medium',
            'points': 1
        }
    
    def _parse_true_false_question(self, content: str) -> Dict[str, Any]:
        """解析判断题"""
        lines = content.split('\n')
        title = lines[0].strip()
        
        is_true = None
        for line in lines[1:]:
            if '答案' in line:
                if re.search(r'[对正是√]', line):
                    is_true = True
                elif re.search(r'[错误否×]', line):
                    is_true = False
        
        return {
            'type': 'true_false',
            'title': title,
            'content': {},
            'answer': {
                'is_true': is_true if is_true is not None else True
            },
            'difficulty': 'medium',
            'points': 1
        }
    
    def _parse_qa_question(self, content: str) -> Dict[str, Any]:
        """解析问答题"""
        lines = content.split('\n')
        title = lines[0].strip()
        
        answer_keywords = []
        for line in lines[1:]:
            if '答案' in line or '参考答案' in line:
                answer_text = re.sub(r'答案[：:]?\s*', '', line).strip()
                if answer_text:
                    answer_keywords.append(answer_text)
        
        return {
            'type': 'qa',
            'title': title,
            'content': {},
            'answer': {
                'keywords': answer_keywords or [title]  # 如果没有找到答案，使用题目作为关键词
            },
            'difficulty': 'medium',
            'points': 2
        }
    
    def _convert_excel_row_to_question(self, row_data: Dict[str, Any]) -> Dict[str, Any]:
        """将Excel行数据转换为题目格式"""
        # 支持中英文表头
        title = (row_data.get('题目') or row_data.get('题干') or
                row_data.get('title') or row_data.get('question'))
        question_type_raw = (row_data.get('类型') or row_data.get('题型') or
                             row_data.get('type') or 'choice')

        # 映射中文题型到英文
        type_mapping = {
            '单选题': 'choice',
            '多选题': 'choice',
            '判断题': 'true_false',
            '问答题': 'qa',
            '填空题': 'qa',
            '计算题': 'math',
            '编程题': 'programming'
        }
        question_type = type_mapping.get(question_type_raw, question_type_raw)
        difficulty = row_data.get('难度') or row_data.get('difficulty') or 'medium'

        if not title:
            return None
        
        question = {
            'type': question_type,
            'title': str(title).strip(),
            'difficulty': difficulty,
            'points': int(row_data.get('分值') or row_data.get('points') or 1)
        }
        
        # 根据题型处理内容和答案
        if question_type == 'choice':
            options = []
            for key in ['A', 'B', 'C', 'D', 'E', 'F']:
                option_text = (row_data.get(f'选项{key}') or
                              row_data.get(f'option_{key}') or
                              row_data.get(f'Option {key}'))
                if option_text and str(option_text).strip():
                    options.append({'key': key, 'text': str(option_text).strip()})

            question['content'] = {'options': options}
            correct_answer = (row_data.get('答案') or row_data.get('正确答案') or
                             row_data.get('answer') or 'A')
            question['answer'] = {'correct_option': str(correct_answer).strip()}
            
        elif question_type == 'true_false':
            question['content'] = {}
            answer = row_data.get('答案') or row_data.get('answer')
            is_true = str(answer).strip() in ['对', '正确', '是', 'True', 'true', '√']
            question['answer'] = {'is_true': is_true}
            
        else:  # qa, math, programming
            question['content'] = {}
            answer = row_data.get('答案') or row_data.get('answer') or ''
            question['answer'] = {'keywords': [str(answer).strip()]}
        
        # 添加解析
        explanation = row_data.get('解析') or row_data.get('explanation')
        if explanation and str(explanation).strip():
            question['explanation'] = str(explanation).strip()
        
        return question
    
    def _validate_question_data(self, question: Dict[str, Any]) -> Dict[str, Any]:
        """验证题目数据格式"""
        required_fields = ['type', 'title', 'content', 'answer']
        
        for field in required_fields:
            if field not in question:
                raise ValueError(f"缺少必需字段: {field}")
        
        # 验证题目类型
        valid_types = ['choice', 'true_false', 'qa', 'math', 'programming']
        if question['type'] not in valid_types:
            raise ValueError(f"无效的题目类型: {question['type']}")
        
        # 设置默认值
        question.setdefault('difficulty', 'medium')
        question.setdefault('points', 1)
        question.setdefault('tags', [])
        
        return question
    
    def import_questions(self, questions_data: List[Dict[str, Any]], bank_id: int) -> int:
        """将题目数据导入到数据库"""
        imported_count = 0
        
        for question_data in questions_data:
            try:
                question_data['bank_id'] = bank_id
                question = Question.create_from_dict(question_data)
                db.session.add(question)
                imported_count += 1
            except Exception as e:
                print(f"导入题目失败: {e}")
                continue
        
        db.session.commit()
        return imported_count
